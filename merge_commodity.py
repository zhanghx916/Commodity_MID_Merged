"""
Commodity_MID_ Excel 文件合并脚本
- 合并所有 Commodity_MID_*.xlsx 文件的 Sheet2
- 按 "Part Description" + "HTS Number" 去重（保留首次出现）
- 完整保留原始格式（字体、填充、边框、对齐、行高、列宽等）
- 字段校验 sheet 直接复制第一个文件的，不做改动
- 输出：Commodity_MID_merged.xlsx

用法：
  python merge_commodity.py
"""

import sys
import glob
import copy
import os

# 修复 Windows 控制台中文编码问题
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────────

def copy_cell(src, dst):
    """把 src 单元格的值和所有样式复制到 dst 单元格"""
    dst.value = src.value
    if src.has_style:
        try:
            dst.font = copy.copy(src.font)
        except Exception:
            pass
        try:
            dst.fill = copy.copy(src.fill)
        except Exception:
            pass
        try:
            dst.alignment = copy.copy(src.alignment)
        except Exception:
            pass
        try:
            dst.border = copy.copy(src.border)
        except Exception:
            pass
        try:
            dst.number_format = src.number_format
        except Exception:
            pass
        try:
            dst.protection = copy.copy(src.protection)
        except Exception:
            pass


def copy_sheet_full(src_ws, dst_ws):
    """完整复制一个 sheet（值 + 样式 + 行高 + 列宽 + 合并单元格）到另一个 sheet"""
    # 复制单元格
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            copy_cell(cell, dst_cell)

    # 复制行高
    for row_idx, rd in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = rd.height
        dst_ws.row_dimensions[row_idx].hidden = rd.hidden

    # 复制列宽
    for col_letter, cd in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = cd.width
        dst_ws.column_dimensions[col_letter].hidden = cd.hidden

    # 复制合并单元格
    for merge in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merge))

    # 复制冻结窗格
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    # 复制标签颜色
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = copy.copy(src_ws.sheet_properties.tabColor)


# ──────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────

def main():
    work_dir = r"C:\Commodity_MID_"
    pattern = os.path.join(work_dir, "Commodity_MID_*.xlsx")
    output_path = os.path.join(work_dir, "Commodity_MID_merged.xlsx")

    # 1. 收集文件（排除已有的合并文件和测试文件，避免重复处理）
    files = sorted(glob.glob(pattern))
    files = [f for f in files if not os.path.basename(f).startswith("Commodity_MID_merged")
             and "test" not in os.path.basename(f).lower()]
    if not files:
        print("未找到任何 Commodity_MID_*.xlsx 文件，退出。")
        sys.exit(1)

    print(f"找到 {len(files)} 个文件：")
    for f in files:
        print(f"  {os.path.basename(f)}")

    # 2. 以第一个文件为模板，读取结构
    print(f"\n读取模板文件：{os.path.basename(files[0])}")
    template_wb = openpyxl.load_workbook(files[0], data_only=True)

    sheet_names = template_wb.sheetnames
    print(f"  Sheet 列表：{sheet_names}")

    # 确定 Sheet2 和字段校验 sheet 的名称
    sheet2_name = "Sheet2"
    jiaoyan_name = sheet_names[1] if len(sheet_names) > 1 else None

    if sheet2_name not in sheet_names:
        print(f"错误：未找到 Sheet2，实际 sheets：{sheet_names}")
        sys.exit(1)

    template_sheet2 = template_wb[sheet2_name]

    # 3. 读取 Sheet2 的标题行，找到 Part Description 和 HTS Number 的列索引
    header_col_map = {}  # header_name.strip() -> col_index (1-based)
    for cell in template_sheet2[1]:
        header_name = str(cell.value).strip() if cell.value is not None else ""
        if header_name:
            header_col_map[header_name] = cell.column

    print(f"\n列标题（共 {len(header_col_map)} 个）：")
    for name, col in sorted(header_col_map.items(), key=lambda x: x[1]):
        print(f"  列{col}({get_column_letter(col)}): '{name}'")

    part_desc_col = header_col_map.get("Part Description")
    hts_col = header_col_map.get("HTS Number")

    if part_desc_col is None or hts_col is None:
        print(f"错误：未找到去重列！")
        print(f"  'Part Description' -> {part_desc_col}")
        print(f"  'HTS Number' -> {hts_col}")
        sys.exit(1)

    print(f"\n去重列：")
    print(f"  Part Description -> 列{part_desc_col} ({get_column_letter(part_desc_col)})")
    print(f"  HTS Number       -> 列{hts_col} ({get_column_letter(hts_col)})")

    # 4. 遍历所有文件，收集去重后的数据行（保持 wb 打开以保留单元格引用）
    # 策略：把每行的数据（值+样式信息）存为字典列表，而不是直接引用单元格对象
    # 这样可以安全关闭 wb 后仍能访问数据

    seen_keys = set()
    # dedup_rows: list of list of (col_index, value, font, fill, alignment, border, number_format)
    dedup_rows = []
    total_before = 0

    for filepath in files:
        fname = os.path.basename(filepath)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheet2_name not in wb.sheetnames:
            print(f"  警告：{fname} 中未找到 Sheet2，跳过")
            wb.close()
            continue

        ws = wb[sheet2_name]
        file_rows = 0
        file_new = 0

        # 从第2行开始（第1行是标题）
        for row in ws.iter_rows(min_row=2):
            # 跳过完全空行
            if all(cell.value is None for cell in row):
                continue

            file_rows += 1
            total_before += 1

            # 计算去重键
            part_val = ""
            hts_val = ""
            for cell in row:
                if cell.column == part_desc_col:
                    part_val = str(cell.value).strip() if cell.value is not None else ""
                elif cell.column == hts_col:
                    hts_val = str(cell.value).strip() if cell.value is not None else ""
            key = (part_val, hts_val)

            if key not in seen_keys:
                seen_keys.add(key)
                file_new += 1
                # 把这行的所有单元格数据序列化保存
                row_data = []
                for cell in row:
                    cell_info = {
                        "col": cell.column,
                        "value": cell.value,
                        "number_format": cell.number_format if cell.has_style else "General",
                    }
                    if cell.has_style:
                        try:
                            cell_info["font"] = copy.copy(cell.font)
                        except Exception:
                            cell_info["font"] = None
                        try:
                            cell_info["fill"] = copy.copy(cell.fill)
                        except Exception:
                            cell_info["fill"] = None
                        try:
                            cell_info["alignment"] = copy.copy(cell.alignment)
                        except Exception:
                            cell_info["alignment"] = None
                        try:
                            cell_info["border"] = copy.copy(cell.border)
                        except Exception:
                            cell_info["border"] = None
                    else:
                        cell_info["font"] = None
                        cell_info["fill"] = None
                        cell_info["alignment"] = None
                        cell_info["border"] = None
                    row_data.append(cell_info)
                dedup_rows.append(row_data)

        print(f"  {fname}: {file_rows} 行，新增 {file_new} 行（去重后）")
        wb.close()

    total_after = len(dedup_rows)
    print(f"\n汇总：原始共 {total_before} 行，去重后 {total_after} 行（去除重复 {total_before - total_after} 行）")

    # 5. 创建输出 workbook
    print(f"\n构建输出文件...")
    out_wb = openpyxl.Workbook()

    # 删除默认的 Sheet
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    # 6. 创建 Sheet2 并写入数据
    out_ws2 = out_wb.create_sheet(sheet2_name)

    # 复制标题行（第1行）的值和样式
    for cell in template_sheet2[1]:
        dst_cell = out_ws2.cell(row=1, column=cell.column)
        copy_cell(cell, dst_cell)

    # 复制标题行高
    if 1 in template_sheet2.row_dimensions:
        out_ws2.row_dimensions[1].height = template_sheet2.row_dimensions[1].height

    # 写入去重后的数据行
    for data_row_idx, row_data in enumerate(dedup_rows, start=2):
        for cell_info in row_data:
            dst_cell = out_ws2.cell(row=data_row_idx, column=cell_info["col"])
            dst_cell.value = cell_info["value"]
            if cell_info.get("number_format"):
                try:
                    dst_cell.number_format = cell_info["number_format"]
                except Exception:
                    pass
            if cell_info.get("font"):
                try:
                    dst_cell.font = cell_info["font"]
                except Exception:
                    pass
            if cell_info.get("fill"):
                try:
                    dst_cell.fill = cell_info["fill"]
                except Exception:
                    pass
            if cell_info.get("alignment"):
                try:
                    dst_cell.alignment = cell_info["alignment"]
                except Exception:
                    pass
            if cell_info.get("border"):
                try:
                    dst_cell.border = cell_info["border"]
                except Exception:
                    pass

        if data_row_idx % 500 == 0:
            print(f"  已写入 {data_row_idx - 1} 行...")

    # 复制列宽（来自模板文件）
    for col_letter, cd in template_sheet2.column_dimensions.items():
        out_ws2.column_dimensions[col_letter].width = cd.width
        out_ws2.column_dimensions[col_letter].hidden = cd.hidden

    # 复制冻结窗格
    if template_sheet2.freeze_panes:
        out_ws2.freeze_panes = template_sheet2.freeze_panes

    # 复制标签颜色
    if template_sheet2.sheet_properties.tabColor:
        out_ws2.sheet_properties.tabColor = copy.copy(template_sheet2.sheet_properties.tabColor)

    print(f"  Sheet2 写入完成：标题1行 + 数据{total_after}行")

    # 7. 复制字段校验 sheet（直接完整复制第一个文件的）
    if jiaoyan_name and jiaoyan_name in template_wb.sheetnames:
        src_jiaoyan = template_wb[jiaoyan_name]
        out_jiaoyan = out_wb.create_sheet(jiaoyan_name)
        copy_sheet_full(src_jiaoyan, out_jiaoyan)
        print(f"  字段校验 sheet 复制完成")
    else:
        print(f"  警告：未找到字段校验 sheet，跳过")

    # 8. 保存
    out_wb.save(output_path)
    print(f"\n合并完成！输出文件：{output_path}")
    print(f"   Sheet2：{total_after + 1} 行（含标题行）")

    template_wb.close()


if __name__ == "__main__":
    main()
